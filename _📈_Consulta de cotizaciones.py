import streamlit as st
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import date, timedelta, datetime
from io import StringIO, BytesIO
from pathlib import Path
import os
import json
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import openpyxl
import urllib3

# Desactivar advertencias de SSL inseguro para la API de GGSA
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- CONFIGURACIÓN DE CONSTANTES Y CONEXIÓN ---
URL = "https://portalerrepar.errepar.com/CotizacionDolarPage"
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Referer': URL
}

CONFIG_FILE = "config.json"

# Configuración de Hojas y Columnas
EXCEL_SHEET = "Divisa-Billete"
EXCEL_COLUMNS = ['Fecha', 'Billete Compra', 'Billete Venta', 'Divisa Compra', 'Divisa Venta']

MEP_SHEET = "MEP"
MEP_COLUMNS = ['fecha', 'DOLAR MEP']
AMBITO_MEP_URL_TMPL = "https://mercados.ambito.com//dolarrava/mep/grafico/{desde}/{hasta}"

LIBRE_SHEET = "Libre"
LIBRE_COLUMNS = ['Fecha', 'Compra', 'Venta']
AMBITO_LIBRE_URL_TMPL = "https://mercados.ambito.com//dolar/informal/historico-general/{desde}/{hasta}"

UVA_SHEET = "UVA"
UVA_COLUMNS = ["Fecha", "Valor"]
UVA_URL = "https://api.argentinadatos.com/v1/finanzas/indices/uva/"

CAC_SHEET = "CAC"
CAC_COLUMNS = ["Periodo", "General", "Materiales", "Mano de obra"]
CAC_URL = "https://prestamos.ikiwi.net.ar/api/cacs"

SMVYM_SHEET = "SMVYM"
SMVYM_COLUMNS = ["Periodo", "Salario"]
SMVYM_URL = "https://apis.datos.gob.ar/series/api/series/?metadata=full&ids=57.1_SMVMM_0_M_34&limit=5000&start=0"

IPC_SHEET = "IPC"
IPC_COLUMNS = ["Fecha", "Valor"]
IPC_URL = "https://apis.datos.gob.ar/series/api/series/?metadata=full&ids=145.3_INGNACNAL_DICI_M_15&limit=5000&start=0"

ROSARIO_SHEET = "Pizarra Rosario"
ROSARIO_START_DATE = "2024-01-01" 
ROSARIO_URL_TMPL = "https://www.ggsa.com.ar/get_pizarra/pros59/2024-01-01/{hasta}/"
ROSARIO_COLUMNS_ORDER = ["Fecha", "Trigo", "Maíz", "Sorgo", "Girasol", "Soja"]
ROSARIO_MAP = {
    "trigo": "Trigo",
    "maiz": "Maíz",
    "sorgo": "Sorgo",
    "girasol": "Girasol",
    "soja": "Soja"
}

INTERVALO_DE_ACTUALIZACION = timedelta(hours=24)

FIXED_PAYLOAD = {
    'ctl00$ScriptManager1': 'ctl00$ContentPlaceHolder1$updPnl|ctl00$ContentPlaceHolder1$btnBuscar',
    '__EVENTTARGET': '',
    '__EVENTARGUMENT': '',
    '__ASYNCPOST': 'true',
    'ctl00$ContentPlaceHolder1$btnBuscar': 'VER DATOS'
}

POST_HEADERS = HEADERS.copy()
POST_HEADERS.update({
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'X-Requested-With': 'XMLHttpRequest'
})


# --- GESTIÓN DE CONFIGURACIÓN Y RUTAS ---

def cargar_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def guardar_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)

def _excel_path() -> Path:
    config = cargar_config()
    ruta = config.get('excel_path', '')
    return Path(ruta)

def is_file_locked(filepath: Path) -> bool:
    """Verifica si el archivo está abierto o bloqueado por otro proceso."""
    if not filepath.exists():
        return False
    try:
        # Intentamos abrir el archivo en modo 'append' (agregar)
        with open(filepath, 'a'):
            pass
        return False
    except IOError:
        # Si lanza un error IOError / PermissionError, el archivo está bloqueado
        return True

def asegurar_hojas_existen(filepath: Path):
    """Verifica si las hojas necesarias existen en el Excel, y las crea si no están."""
    hojas_necesarias = [
        EXCEL_SHEET, MEP_SHEET, LIBRE_SHEET, UVA_SHEET, 
        CAC_SHEET, SMVYM_SHEET, IPC_SHEET, ROSARIO_SHEET
    ]
    try:
        wb = openpyxl.load_workbook(filepath)
        modificado = False
        for hoja in hojas_necesarias:
            if hoja not in wb.sheetnames:
                wb.create_sheet(hoja)
                modificado = True
        
        # Opcional: Si el archivo es nuevo y tiene la hoja por defecto vacía, la eliminamos
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
            modificado = True

        if modificado:
            wb.save(filepath)
            st.success("✅ Se verificaron y crearon las hojas faltantes en el Excel.")
    except Exception as e:
        st.error(f"❌ Error al verificar/crear hojas: {e}")


# --- FUNCIONES DE SCRAPING (BNA) ---

@st.cache_data(ttl=3600)
def get_dynamic_payload_fields(_session):
    st.info("🔄 Obteniendo estado de página (ViewState) de Errepar...")
    try:
        initial_response = _session.get(URL, headers=HEADERS)
        initial_response.raise_for_status()
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Error de conexión inicial con Errepar: {e}")
        return None

    soup = BeautifulSoup(initial_response.text, 'html.parser')
    payload_fields = {}
    try:
        payload_fields['__VIEWSTATE'] = soup.find('input', {'id': '__VIEWSTATE'})['value']
        payload_fields['__VIEWSTATEGENERATOR'] = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})['value']
        payload_fields['__EVENTVALIDATION'] = soup.find('input', {'id': '__EVENTVALIDATION'})['value']
    except (TypeError, KeyError):
        st.error("❌ Fallo al encontrar campos de estado en Errepar.")
        return None

    payload_fields.update(FIXED_PAYLOAD)
    return payload_fields

def obtener_cotizaciones(fecha_desde: str, fecha_hasta: str) -> pd.DataFrame:
    with requests.Session() as session:
        base_payload = get_dynamic_payload_fields(_session=session)
        if base_payload is None:
            return pd.DataFrame()

        st.info(f"🔎 Solicitando datos de BNA desde **{fecha_desde}** hasta **{fecha_hasta}**...")
        payload = base_payload.copy()
        payload['ctl00$ContentPlaceHolder1$inputDateDesde'] = fecha_desde
        payload['ctl00$ContentPlaceHolder1$inputDateHasta'] = fecha_hasta

        try:
            response = session.post(URL, data=payload, headers=POST_HEADERS)
            response.raise_for_status()
            update_panel_marker = 'updatePanel|ContentPlaceHolder1_updPnl|'
            if update_panel_marker not in response.text:
                return pd.DataFrame()
            
            html_start = response.text.find(update_panel_marker) + len(update_panel_marker)
            html_end = response.text.find('|0|hiddenField|__EVENTTARGET', html_start)
            if html_end == -1:
                html_end = response.text.find('|7310|scriptStartupBlock', html_start)
            
            html_fragment = response.text[html_start:html_end].strip() if html_end != -1 else response.text[html_start:].strip()
            soup = BeautifulSoup(html_fragment, 'html.parser')
            table = soup.find('table', class_='table')
            
            if table is None:
                return pd.DataFrame()
                
            data = []
            rows = table.find_all('tr')
            for row in rows[2:-1]:
                cols = row.find_all('td')
                if len(cols) == 5:
                    data.append([col.text.strip() for col in cols])
            
            if not data:
                return pd.DataFrame()
                
            df = pd.DataFrame(data, columns=EXCEL_COLUMNS)
            for col in ['Billete Compra', 'Billete Venta', 'Divisa Compra', 'Divisa Venta']:
                df[col] = pd.to_numeric(df[col].str.replace(',', '.'))
            
            return normalizar_formato_fecha(df, 'Fecha', dayfirst=True)

        except Exception as e:
            st.error(f"❌ Error BNA: {e}")
            return pd.DataFrame()

# --- UTILIDADES PARA EXCEL ---

def normalizar_formato_fecha(df: pd.DataFrame, col_name: str = 'Fecha', dayfirst: bool = True) -> pd.DataFrame:
    df = df.copy()
    df[col_name] = pd.to_datetime(
        df[col_name], 
        dayfirst=dayfirst, 
        errors='coerce'
    ).dt.date
    return df

def leer_ultima_fecha_excel(sheet_name=EXCEL_SHEET, date_col='Fecha') -> date | None:
    try:
        path = _excel_path()
        if not path.exists(): return None
        df = pd.read_excel(path, sheet_name=sheet_name)
        if date_col not in df.columns: return None
        fechas = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
        return fechas.max().date() if not pd.isna(fechas.max()) else None
    except Exception: return None

def actualizar_hoja_excel(df_nuevo: pd.DataFrame, sheet_name: str, key_column: str, format_cols: dict = None) -> bool:
    try:
        path = _excel_path()
        if not path.exists():
            st.error(f"❌ El archivo Excel no existe en: {path}")
            return False
            
        try:
            df_existente = pd.read_excel(path, sheet_name=sheet_name)
            if sheet_name in [SMVYM_SHEET, UVA_SHEET, IPC_SHEET]:
                df_existente = normalizar_formato_fecha(df_existente, key_column, dayfirst=False)
            else:
                df_existente = normalizar_formato_fecha(df_existente, key_column, dayfirst=True)
        except Exception:
            df_existente = pd.DataFrame()

        df_total = pd.concat([df_existente, df_nuevo], ignore_index=True)
        if sheet_name == CAC_SHEET:
            df_total[key_column] = df_total[key_column].astype(str)
            
        df_total = df_total.drop_duplicates(subset=[key_column], keep='last').sort_values(key_column)

        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_total.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            
            ws.freeze_panes = 'A2'
            for column_cells in ws.columns:
                col_letter = column_cells[0].column_letter
                ws.column_dimensions[col_letter].width = 12
            
            if format_cols:
                for col_name, num_format in format_cols.items():
                    if col_name in df_total.columns:
                        col_idx = df_total.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        for cell in ws[col_letter][1:]:
                            cell.number_format = num_format

        st.success(f"💾 Excel actualizado: hoja '{sheet_name}'.")
        return True
    except Exception as e:
        st.error(f"❌ Error actualizando hoja '{sheet_name}': {e}")
        return False

# --- APIS GENERALES ---

def obtener_datos_api(url: str, source_name: str, verify: bool = True) -> dict | list | None:
    st.info(f"🔎 Consultando API de {source_name}...")
    try:
        r = requests.get(url, headers={'User-Agent': HEADERS['User-Agent']}, timeout=30, verify=verify)
        r.raise_for_status()
        st.success(f"✅ Respuesta de API {source_name} recibida.")
        return r.json()
    except Exception as e:
        st.error(f"❌ Error {source_name}: {e}")
        return None

def obtener_mep(desde: date, hasta: date) -> pd.DataFrame:
    url = AMBITO_MEP_URL_TMPL.format(desde=desde.strftime('%Y-%m-%d'), hasta=hasta.strftime('%Y-%m-%d'))
    data = obtener_datos_api(url, "MEP")
    if not isinstance(data, list) or len(data) < 2: return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=MEP_COLUMNS)
    df['DOLAR MEP'] = pd.to_numeric(df['DOLAR MEP'].astype(str).str.replace(',', '.', regex=False), errors='coerce')
    return normalizar_formato_fecha(df, 'fecha', dayfirst=True).dropna(subset=['fecha'])

def obtener_libre(desde: date, hasta: date) -> pd.DataFrame:
    url = AMBITO_LIBRE_URL_TMPL.format(desde=desde.strftime('%Y-%m-%d'), hasta=hasta.strftime('%Y-%m-%d'))
    data = obtener_datos_api(url, "LIBRE")
    if not isinstance(data, list) or len(data) < 2: return pd.DataFrame()
    df = pd.DataFrame(data[1:], columns=LIBRE_COLUMNS)
    for col in ['Compra', 'Venta']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.', regex=False), errors='coerce')
    return normalizar_formato_fecha(df, dayfirst=True).dropna(subset=['Fecha'])

def obtener_uva() -> pd.DataFrame:
    data = obtener_datos_api(UVA_URL, "UVA")
    if not isinstance(data, list): return pd.DataFrame()
    df = pd.DataFrame(data)
    df.rename(columns={'fecha': 'Fecha', 'valor': 'Valor'}, inplace=True)
    return normalizar_formato_fecha(df, dayfirst=False).dropna(subset=['Fecha'])

def obtener_cac() -> pd.DataFrame:
    data = obtener_datos_api(CAC_URL, "CAC")
    if not isinstance(data, list): return pd.DataFrame()
    df = pd.DataFrame(data)
    df.rename(columns={'period': 'Periodo', 'general': 'General', 'materials': 'Materiales', 'labour_force': 'Mano de obra'}, inplace=True)
    df = df[['Periodo', 'General', 'Materiales', 'Mano de obra']]
    for col in ['General', 'Materiales', 'Mano de obra']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')
    df['Periodo'] = pd.to_datetime(df['Periodo']).dt.strftime('%Y-%m-%d')
    return df.dropna(subset=['Periodo'])

def obtener_smvym() -> pd.DataFrame:
    data = obtener_datos_api(SMVYM_URL, "SMVyM")
    if not isinstance(data, dict) or 'data' not in data: return pd.DataFrame()
    df = pd.DataFrame(data['data'], columns=['Periodo', 'Salario'])
    df['Periodo'] = pd.to_datetime(df['Periodo']).dt.date
    return df.dropna(subset=['Periodo'])

def obtener_ipc() -> pd.DataFrame:
    data = obtener_datos_api(IPC_URL, "IPC")
    if not isinstance(data, dict) or 'data' not in data: return pd.DataFrame()
    df = pd.DataFrame(data['data'], columns=['Periodo', 'Valor'])
    df.rename(columns={'Periodo': 'Fecha'}, inplace=True) 
    df['Fecha'] = pd.to_datetime(df['Fecha']).dt.date
    return df.dropna(subset=['Fecha'])

# --- PIZARRA ROSARIO (GGSA) ---

def obtener_datos_rosario(fecha_hasta: date) -> pd.DataFrame:
    url = ROSARIO_URL_TMPL.format(hasta=fecha_hasta.strftime('%Y-%m-%d'))
    json_data = obtener_datos_api(url, "Pizarra Rosario (GGSA)", verify=False)
    
    if not json_data or "pizarra" not in json_data:
        df = pd.DataFrame(columns=["Fecha"] + list(ROSARIO_MAP.values()))
    else:
        pizarra_dict = json_data["pizarra"]
        rows = []

        for fecha_str, items in pizarra_dict.items():
            row = {"Fecha": fecha_str}
            for api_key, excel_col in ROSARIO_MAP.items():
                if api_key in items:
                    try:
                        precio = float(items[api_key].get("precio", 0))
                        estimativo = float(items[api_key].get("estimativo", 0))
                        
                        valor_final = 0.0
                        es_estimativo = False
                        
                        if precio > 0:
                            valor_final = precio
                            es_estimativo = False
                        elif estimativo > 0:
                            valor_final = estimativo
                            es_estimativo = True
                        
                        row[excel_col] = valor_final
                        row[f"{excel_col}_is_est"] = es_estimativo
                        
                    except (ValueError, TypeError):
                        row[excel_col] = 0.0
                        row[f"{excel_col}_is_est"] = False
                else:
                    row[excel_col] = 0.0
                    row[f"{excel_col}_is_est"] = False
            rows.append(row)
        
        df = pd.DataFrame(rows)

    if not df.empty and "Fecha" in df.columns:
         df["Fecha"] = pd.to_datetime(df["Fecha"])

    try:
        fecha_inicio_dt = pd.to_datetime(ROSARIO_START_DATE)
        fecha_fin_dt = pd.to_datetime(fecha_hasta)
        
        rango_fechas = pd.date_range(start=fecha_inicio_dt, end=fecha_fin_dt, freq='D')
        
        if not df.empty:
            df = df.set_index("Fecha")
        
        df = df.reindex(rango_fechas)
        df = df.ffill()
        df = df.reset_index().rename(columns={"index": "Fecha"})
        df = df.fillna(0)
    except Exception as e:
        st.warning(f"⚠️ Advertencia procesando relleno fechas Rosario: {e}")

    if not df.empty and "Fecha" in df.columns:
         df["Fecha"] = pd.to_datetime(df["Fecha"]).dt.date
    
    return df.sort_values("Fecha")

def guardar_rosario_con_estilo(df: pd.DataFrame) -> bool:
    sheet_name = ROSARIO_SHEET
    path = _excel_path()
    
    if not path.exists():
        st.error("No existe el archivo Excel base.")
        return False
        
    st.info(f"🎨 Guardando y aplicando estilos en hoja '{sheet_name}'...")
    
    try:
        cols_valores = ["Fecha"] + list(ROSARIO_MAP.values())
        cols_estilos = {col: f"{col}_is_est" for col in ROSARIO_MAP.values()}
        
        for col in cols_valores:
             if col not in df.columns: df[col] = 0
        
        df_to_write = df[cols_valores].copy()
        
        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            
            ws.freeze_panes = 'A2'
            for column_cells in ws.columns:
                col_letter = column_cells[0].column_letter
                ws.column_dimensions[col_letter].width = 12
            
            fill_celeste = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            num_format = '#,##0.00'
            
            for index, row in df.iterrows():
                excel_row = index + 2
                for col_name, col_bool_name in cols_estilos.items():
                    if col_name in df_to_write.columns:
                        col_idx = df_to_write.columns.get_loc(col_name) + 1
                        cell = ws.cell(row=excel_row, column=col_idx)
                        cell.number_format = num_format
                        
                        if col_bool_name in df.columns and row.get(col_bool_name, False):
                            cell.fill = fill_celeste

        st.success(f"✅ Hoja '{sheet_name}' guardada con formato condicional (celeste para estimativos).")
        return True

    except Exception as e:
        st.error(f"❌ Error guardando hoja Rosario: {e}")
        return False


# --- FUNCIONES DE RELLENO Y POST-PROCESO ---

def rellenar_fechas_faltantes(df: pd.DataFrame, key_column: str, exclude_dates: list[date]) -> pd.DataFrame:
    if df.empty: return pd.DataFrame()
    df_temp = df.copy()
    original_dates = set(df_temp[key_column])
    df_temp[key_column] = pd.to_datetime(df_temp[key_column])
    df_temp = df_temp.set_index(key_column).sort_index()
    
    start_date = df_temp.index.min()
    end_date = df_temp.index.max()
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    
    df_filled = df_temp.reindex(date_range).ffill()
    df_filled = df_filled.reset_index().rename(columns={'index': key_column})
    df_filled[key_column] = df_filled[key_column].dt.date
    
    df_final = df_filled.copy()
    for exclude_date in exclude_dates:
        if exclude_date in df_final[key_column].values and exclude_date not in original_dates:
            df_final = df_final[df_final[key_column] != exclude_date]

    return df_final.sort_values(key_column).reset_index(drop=True)

def post_process_and_fill_sheet(sheet_name: str, key_column: str, hoy: date) -> bool:
    st.info(f"🔧 Post-Proceso: Rellenando fechas faltantes en hoja '{sheet_name}'...")
    try:
        path = _excel_path()
        if not path.exists(): return False

        df_sheet = pd.read_excel(path, sheet_name=sheet_name)
        
        if sheet_name == EXCEL_SHEET:
            format_cols = {col: '#,##0.00' for col in EXCEL_COLUMNS[1:]}
        elif sheet_name == MEP_SHEET:
            format_cols = {'DOLAR MEP': '#,##0.00'}
        elif sheet_name == LIBRE_SHEET:
            format_cols = {'Compra': '#,##0.00', 'Venta': '#,##0.00'}
        elif sheet_name == ROSARIO_SHEET:
             return True
        else:
            return False

        df_sheet = normalizar_formato_fecha(df_sheet, key_column, dayfirst=True) 
        ayer = hoy - timedelta(days=1)
        df_filled = rellenar_fechas_faltantes(df_sheet, key_column, [hoy, ayer])
        
        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_filled.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            
            ws.freeze_panes = 'A2'
            for column_cells in ws.columns:
                col_letter = column_cells[0].column_letter
                ws.column_dimensions[col_letter].width = 12
            
            for col_name, num_format in format_cols.items():
                if col_name in df_filled.columns:
                    col_idx = df_filled.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for cell in ws[col_letter][1:]:
                        cell.number_format = num_format
            
        st.success(f"✅ Hoja '{sheet_name}' rellenada.")
        return True
    except Exception as e:
        st.error(f"❌ Error Post-Proceso '{sheet_name}': {e}")
        return False

# --- PROCESO PRINCIPAL ---

def ejecutar_proceso_completo_de_actualizacion() -> bool:
    st.write("---")
    st.info("Iniciando proceso de actualización de datos...")
    
    ruta_excel = _excel_path()
    
    # 0. VERIFICAR QUE EL ARCHIVO NO ESTÉ BLOQUEADO/ABIERTO
    if is_file_locked(ruta_excel):
        st.error("❌ El archivo Excel se encuentra abierto o bloqueado por otro programa. Por favor, ciérralo y vuelve a intentarlo.")
        return False

    # 1. ASEGURAR QUE LAS HOJAS EXISTEN ANTES DE LEER
    asegurar_hojas_existen(ruta_excel)
    
    hoy = date.today()
    ayer = hoy - timedelta(days=1)

    # 2. ACTUALIZAR DIVISAS
    ultima_bna = leer_ultima_fecha_excel(EXCEL_SHEET, 'Fecha')
    ultima_mep = leer_ultima_fecha_excel(MEP_SHEET, 'fecha')
    ultima_libre = leer_ultima_fecha_excel(LIBRE_SHEET, 'Fecha')
    
    desde_bna = (ultima_bna + timedelta(days=1)) if ultima_bna else date(2023, 1, 1)
    desde_mep = (ultima_mep + timedelta(days=1)) if ultima_mep else date(2023, 1, 1)
    desde_libre = (ultima_libre + timedelta(days=1)) if ultima_libre else date(2023, 1, 1)
    
    if not ultima_bna or desde_bna <= ayer:
        df_bna = obtener_cotizaciones(desde_bna.strftime("%d/%m/%Y"), ayer.strftime("%d/%m/%Y"))
        if not df_bna.empty:
            actualizar_hoja_excel(df_bna, EXCEL_SHEET, 'Fecha', {col: '#,##0.00' for col in EXCEL_COLUMNS[1:]})

    if not ultima_mep or desde_mep <= ayer:
        df_mep = obtener_mep(desde_mep, ayer)
        if not df_mep.empty:
            actualizar_hoja_excel(df_mep, MEP_SHEET, 'fecha', {'DOLAR MEP': '#,##0.00'})

    if not ultima_libre or desde_libre <= hoy:
        df_libre = obtener_libre(desde_libre, hoy)
        if not df_libre.empty:
            actualizar_hoja_excel(df_libre, LIBRE_SHEET, 'Fecha', {'Compra': '#,##0.00', 'Venta': '#,##0.00'})

    # 3. ACTUALIZAR ÍNDICES
    df_uva = obtener_uva()
    if not df_uva.empty:
        actualizar_hoja_excel(df_uva, UVA_SHEET, 'Fecha', {'Valor': '#,##0.00'})
    
    df_cac = obtener_cac()
    if not df_cac.empty:
        actualizar_hoja_excel(df_cac, CAC_SHEET, 'Periodo', {'General': '#,##0.00', 'Materiales': '#,##0.00', 'Mano de obra': '#,##0.00'})

    df_smvym = obtener_smvym()
    if not df_smvym.empty:
        actualizar_hoja_excel(df_smvym, SMVYM_SHEET, 'Periodo', {'Salario': '#,##0.00'})
        
    df_ipc = obtener_ipc()
    if not df_ipc.empty:
        actualizar_hoja_excel(df_ipc, IPC_SHEET, 'Fecha', {'Valor': '#,##0.00'})

    # 4. ACTUALIZAR PIZARRA ROSARIO
    df_rosario = obtener_datos_rosario(ayer)
    if not df_rosario.empty:
        guardar_rosario_con_estilo(df_rosario)

    # 5. POST-PROCESO
    st.divider()
    st.subheader("🛠️ Post-Proceso: Relleno de Fechas")
    post_process_and_fill_sheet(EXCEL_SHEET, 'Fecha', hoy)
    post_process_and_fill_sheet(MEP_SHEET, 'fecha', hoy)
    post_process_and_fill_sheet(LIBRE_SHEET, 'Fecha', hoy)

    # GUARDAR LA FECHA Y HORA DE ACTUALIZACIÓN EN EL DISCO
    config = cargar_config()
    config['last_update'] = datetime.now().isoformat()
    guardar_config(config)

    st.success("✅ Proceso completo finalizado.")
    return True

def ui_configuracion_inicial():
    st.title("⚙️ Configuración Inicial")
    st.warning("No se ha configurado la ruta del archivo Excel. Por favor, selecciona un archivo o una carpeta.")
    
    # Mantenemos la variable en session_state pero ya no atada a un widget directamente
    if 'excel_path_input' not in st.session_state:
        st.session_state.excel_path_input = ""

    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📂 Buscar Archivo Existente", use_container_width=True):
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes('-topmost', 1)
            archivo = filedialog.askopenfilename(
                master=root, 
                title="Seleccionar archivo Excel", 
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            root.destroy()
            if archivo:
                st.session_state.excel_path_input = archivo
                st.rerun()
                
    with col2:
        if st.button("📁 Seleccionar Carpeta (Nuevo Archivo)", use_container_width=True):
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes('-topmost', 1)
            carpeta = filedialog.askdirectory(
                master=root, 
                title="Seleccionar carpeta de destino"
            )
            root.destroy()
            if carpeta:
                st.session_state.excel_path_input = f"{carpeta}/Cotizaciones y datos macro.xlsx"
                st.rerun()

    ruta_actual = st.session_state.excel_path_input
    
    if ruta_actual:
        st.info(f"📍 Ruta seleccionada: **{ruta_actual}**")
        crear_archivo = st.checkbox("Crear archivo automáticamente si la ruta no existe", value=True)
        
        if st.button("Guardar Configuración", type="primary"):
            ruta_obj = Path(ruta_actual)
            
            if not ruta_actual.endswith('.xlsx'):
                st.error("❌ El archivo debe tener la extensión .xlsx")
                return
                
            if ruta_obj.exists() and ruta_obj.is_file():
                config = cargar_config()
                config['excel_path'] = str(ruta_obj)
                guardar_config(config)
                st.success("✅ Ruta guardada correctamente.")
                st.rerun()
                
            elif not ruta_obj.exists() and crear_archivo:
                if ruta_obj.parent.exists():
                    try:
                        wb = openpyxl.Workbook()
                        wb.save(ruta_obj)
                        
                        config = cargar_config()
                        config['excel_path'] = str(ruta_obj)
                        guardar_config(config)
                        
                        st.success("✅ Archivo creado y ruta guardada correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al crear el archivo: {e}")
                else:
                    st.error(f"❌ La carpeta padre no existe: {ruta_obj.parent}. Por favor, usa el botón '📁 Seleccionar Carpeta...' para elegir un directorio válido.")
            else:
                st.error("❌ El archivo no existe y no seleccionaste la opción de crearlo.")

def main():
    st.set_page_config(page_title="Scraper Financiero & Agro", layout="centered")
    
    # Validar que exista la configuración primero
    config = cargar_config()
    if 'excel_path' not in config:
        ui_configuracion_inicial()
        return  # Se detiene aquí hasta que configuren el Excel

    ruta_excel_actual = config['excel_path']

    # --- LÓGICA DE ACTUALIZACIÓN PERSISTENTE (MODIFICADO) ---
    last_update_str = config.get('last_update')
    if last_update_str:
        try:
            last_update = datetime.fromisoformat(last_update_str)
        except ValueError:
            last_update = datetime.min
    else:
        # Si nunca se actualizó, forzamos la fecha mínima
        last_update = datetime.min

    ahora = datetime.now()
    tiempo_pasado = ahora - last_update

    st.title("💸 Actualizador Financiero & Agro")
    
    # Mostrar la ruta actual y dar la opción de cambiarla
    with st.expander(f"📁 Archivo destino: {ruta_excel_actual}", expanded=False):
        if st.button("Cambiar archivo de destino"):
            # Borrar path de config y del session state
            del config['excel_path']
            guardar_config(config)
            if 'excel_path_input' in st.session_state:
                del st.session_state.excel_path_input
            st.rerun()

    # Interfaz visual de la carga
    if tiempo_pasado < INTERVALO_DE_ACTUALIZACION:
        st.info(f"✅ Última actualización completada: **{last_update.strftime('%d/%m/%Y %H:%M:%S')}**")
    else:
        st.warning("⏳ Actualización pendiente (Han pasado más de 24 hs).")

    # Botón manual
    if st.button("🔄 Forzar Actualización Ahora", type="primary", use_container_width=True):
        exito = ejecutar_proceso_completo_de_actualizacion()
        if exito:
            st.rerun() 

    # --- EJECUCIÓN AUTOMÁTICA ---
    # Si pasaron las 24 horas y el usuario abre la página, lo ejecuta solo.
    if tiempo_pasado > INTERVALO_DE_ACTUALIZACION:
        with st.spinner('Actualización automática en curso...'):
            exito = ejecutar_proceso_completo_de_actualizacion()
        # Se recarga la página para mostrar el cartel de éxito y ocultar el spinner (solo si fue exitoso)
        if exito:
            st.rerun()

if __name__ == "__main__":
    main()